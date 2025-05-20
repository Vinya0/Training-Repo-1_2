import openpyxl

# # same data
# file="/home/web-h-044/Downloads/Writing.xlsx"
#
# workbook=openpyxl.load_workbook(file)
# sheet=workbook.active   #  (or) sheet=workbook["Data"]     --  get active sheet from excel
# for r in range(1,7):
#     for c in range(1,5):
#         sheet.cell(r,c).value="welcome"
# workbook.save(file)

#multiple data
file = "/home/web-h-044/Downloads/Writing (1).xlsx"

workbook = openpyxl.load_workbook(file)
sheet = workbook.active  #  (or) sheet=workbook["Data"]     --  get active sheet from excel

sheet.cell(1, 1).value = 123
sheet.cell(1, 2).value = "smith"
sheet.cell(1, 3).value = "engineer"

sheet.cell(2, 1).value = 567
sheet.cell(2, 2).value = "john"
sheet.cell(2, 3).value = "manager"

sheet.cell(3, 1).value = 567
sheet.cell(3, 2).value = "david"
sheet.cell(3, 3).value = "developer"

workbook.save(file)  # save the file after entering the data
