# File--Workbook--Sheets--Rows--Cells
import openpyxl

file = "/home/web-h-044/Downloads/Book (1).xlsx"
workbook=openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]
rows=sheet.max_row #Count number of rows
cols=sheet.max_column #Count number of columns

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end="               ")
    print()